"""
Script pour enrichir le fichier DALKIA.xlsx avec la colonne TYPE
en récupérant la valeur 'categorie_entreprise' depuis l'API recherche-entreprises.api.gouv.fr

Dépendances : pip install openpyxl requests
"""

import time

import openpyxl
import requests

# ── CONFIG ──────────────────────────────────────────────────────────────────
INPUT_FILE = r"c:\Users\C49982\Desktop\Nouveau dossier\DALKIA.xlsx"
OUTPUT_FILE = r"c:\Users\C49982\Desktop\Nouveau dossier\DALKIA_enrichi.xlsx"
API_BASE = "https://recherche-entreprises.api.gouv.fr/search"
DELAY = 0.3
PROXIES = {
    "http": "http://vip-users.proxy.edf.fr:3131",
    "https": "http://vip-users.proxy.edf.fr:3131",
}
# ────────────────────────────────────────────────────────────────────────────


def get_categorie_entreprise(siren: str) -> str:
    siren = str(siren).strip().replace(" ", "").zfill(9)
    try:
        resp = requests.get(API_BASE, params={"q": siren}, timeout=10, proxies=PROXIES)
        resp.raise_for_status()
        data = resp.json()
        results = data.get("results", [])
        if results:
            return results[0].get("categorie_entreprise", "") or ""
    except Exception as e:
        print(f"  Erreur pour SIREN {siren}: {e}")
    return ""


def find_col_index(headers, keywords):
    for i, h in enumerate(headers, start=1):
        if h and any(kw.lower() in str(h).lower() for kw in keywords):
            return i
    return None


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 2)]
    siren_col = find_col_index(headers, ["siren"])
    type_col = find_col_index(headers, ["type"])

    if siren_col is None:
        raise ValueError("Colonne SIREN introuvable.")

    if type_col is None:
        type_col = ws.max_column + 1
        ws.cell(row=1, column=type_col).value = "TYPE"
        print(f"Colonne TYPE créée en colonne {type_col}.")
    else:
        print(f"Colonne TYPE trouvée en colonne {type_col}.")

    total = ws.max_row - 1
    print(f"Traitement de {total} lignes...\n")

    for row in range(2, ws.max_row + 1):
        siren_val = ws.cell(row=row, column=siren_col).value
        if not siren_val:
            continue

        categorie = get_categorie_entreprise(str(siren_val))
        ws.cell(row=row, column=type_col).value = categorie
        print(f"  [{row - 1}/{total}] SIREN {siren_val} -> {categorie or '(vide)'}")
        time.sleep(DELAY)

    wb.save(OUTPUT_FILE)
    print(f"\nFichier sauvegardé : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
